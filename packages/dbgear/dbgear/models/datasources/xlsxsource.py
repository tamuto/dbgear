import openpyxl

from .base import BaseDataSource
from ...utils.dict_utils import dict_to_nested


class DataSource(BaseDataSource):
    folder: str
    environ: str
    name: str
    schema_name: str
    table_name: str
    segment: str | None = None

    def __init__(self, folder: str, data_path: str, table_name: str, header_row: int, start_row: int, **kwargs):
        self.folder = folder
        self.data_path = data_path
        self.table_name = table_name
        self.header_row = header_row
        self.start_row = start_row
        self._data = []

    @property
    def filename(self) -> str:
        return self.data_path

    @property
    def data(self):
        return self._data

    def load(self):
        wb = openpyxl.load_workbook(f'{self.folder}/{self.data_path}', data_only=True)
        ws = wb[self.table_name]

        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=self.header_row, column=col).value
            if cell_value is not None:
                headers.append(str(cell_value).strip())
            else:
                headers.append(f"Column_{col}")

        data = []
        for row_num in range(self.start_row, ws.max_row + 1):
            row_data = {}
            has_data = False

            for col_idx, header in enumerate(headers):
                cell_value = ws.cell(row=row_num, column=col_idx + 1).value

                if cell_value is not None:
                    has_data = True
                    row_data[header] = self._convert_cell_value(cell_value)
                else:
                    row_data[header] = None

            if has_data:
                data.append(dict_to_nested(row_data))

        wb.close()
        self._data = data

    def _convert_cell_value(self, value):
        if value is None:
            return None

        # Handle datetime objects
        if hasattr(value, 'isoformat'):
            return value.isoformat()

        # Handle numeric values
        if isinstance(value, (int, float)):
            # Convert float to int if it's a whole number
            if isinstance(value, float) and value.is_integer():
                return int(value)
            return value

        # Handle string values
        if isinstance(value, str):
            value = value.strip()

            # Handle special values
            if value.upper() in ('NOW()', 'CURRENT_TIMESTAMP'):
                return 'NOW()'
            elif value.upper() in ('SYSTEM', 'CURRENT_USER'):
                return 'SYSTEM'
            elif value.upper() in ('NULL', 'NONE', ''):
                return None

            # Try to convert to number if it looks like one
            try:
                if '.' in value:
                    return float(value)
                else:
                    return int(value)
            except ValueError:
                pass

            return value

        # Return as-is for other types
        return value
