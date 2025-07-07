"""
Excel data importer using openpyxl.
"""

import openpyxl
from typing import Any, Dict, List, Optional
from pathlib import Path

from .base import BaseDataImporter


class ExcelDataImporter(BaseDataImporter):
    """
    Excel data importer using openpyxl library.
    """

    def read_data(self, source: str, **kwargs) -> List[Dict[str, Any]]:
        """
        Read data from Excel file.

        Args:
            source: Path to Excel file
            **kwargs: Additional arguments
                - sheet_name: Name of the sheet to read (default: first sheet)
                - header_row: Row number containing headers (default: 1)
                - start_row: First data row (default: 2)
                - end_row: Last data row (default: None - read all)
                - start_col: First data column (default: 1)
                - end_col: Last data column (default: None - read all)

        Returns:
            List of dictionaries representing rows
        """
        source_path = Path(source)
        if not source_path.exists():
            raise FileNotFoundError(f"Excel file not found: {source}")

        # Load workbook
        workbook = openpyxl.load_workbook(source_path, data_only=True)

        # Get worksheet
        sheet_name = kwargs.get('sheet_name')
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in Excel file")
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active

        # Get parameters
        header_row = kwargs.get('header_row', 1)
        start_row = kwargs.get('start_row', header_row + 1)
        end_row = kwargs.get('end_row', worksheet.max_row)
        start_col = kwargs.get('start_col', 1)
        end_col = kwargs.get('end_col', worksheet.max_column)

        # Read headers
        headers = []
        for col in range(start_col, end_col + 1):
            cell_value = worksheet.cell(row=header_row, column=col).value
            if cell_value is not None:
                headers.append(str(cell_value).strip())
            else:
                headers.append(f"Column_{col}")

        # Read data
        data = []
        for row_num in range(start_row, end_row + 1):
            row_data = {}
            has_data = False

            for col_idx, header in enumerate(headers):
                col_num = start_col + col_idx
                cell_value = worksheet.cell(row=row_num, column=col_num).value

                # Convert cell value to appropriate type
                if cell_value is not None:
                    has_data = True
                    row_data[header] = self._convert_cell_value(cell_value)
                else:
                    row_data[header] = None

            # Only add row if it has at least one non-empty value
            if has_data:
                data.append(row_data)

        workbook.close()
        return data

    def _convert_cell_value(self, value: Any) -> Any:
        """
        Convert Excel cell value to appropriate Python type.

        Args:
            value: Cell value from openpyxl

        Returns:
            Converted value
        """
        if value is None:
            return None

        # Handle datetime objects
        if hasattr(value, 'isoformat'):
            return value.isoformat()

        # Handle numeric values
        if isinstance(value, (int, float)):
            # Convert float to int if it's a whole number
            if isinstance(value, float) and value.is_integer():
                return int(value)
            return value

        # Handle string values
        if isinstance(value, str):
            value = value.strip()

            # Handle special values
            if value.upper() in ('NOW()', 'CURRENT_TIMESTAMP'):
                return 'NOW()'
            elif value.upper() in ('SYSTEM', 'CURRENT_USER'):
                return 'SYSTEM'
            elif value.upper() in ('NULL', 'NONE', ''):
                return None

            # Try to convert to number if it looks like one
            try:
                if '.' in value:
                    return float(value)
                else:
                    return int(value)
            except ValueError:
                pass

            return value

        # Return as-is for other types
        return value


def import_data(source: str, schema, table_name: str, output_path: str,
                schema_name: str = 'main', json_fields: Optional[List[str]] = None,
                **kwargs) -> str:
    """
    Import data from Excel file and write to .dat file.

    Args:
        source: Path to Excel file
        schema: Schema manager containing table definitions
        table_name: Name of the target table
        output_path: Directory path for output file
        schema_name: Name of the schema (default: 'main')
        json_fields: List of field names that should be treated as JSON paths
        **kwargs: Additional arguments for Excel reading

    Returns:
        Path to the created .dat file
    """
    importer = ExcelDataImporter(schema, table_name, schema_name)
    return importer.import_data(source, output_path, json_fields, **kwargs)
